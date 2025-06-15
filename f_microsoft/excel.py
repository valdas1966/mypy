from openpyxl import Workbook, load_workbook, Worksheet
from f_file.i_0_handler import FileHandler


class Excel(FileHandler):
    """
    ============================================================================
     Excel file handler.
    ============================================================================
    """

    def __init__(self,
                 # Path to the Excel file to be opened
                 path: str
                 ) -> None:
        """
        ========================================================================
         Initialize the Excel file handler.
        ========================================================================
        """
        FileHandler.__init__(self, path=path)
        self._wb: Workbook = None
        
    def _create(self) -> None:
        """
        ========================================================================
         Create a new Excel file.
        ========================================================================
        """
        self._wb = Workbook()

    def _open(self) -> None:
        """
        ========================================================================
         Open an existing Excel file.
        ========================================================================
        """
        self._wb = load_workbook(self.path)

    def save(self) -> None:
        """
        ========================================================================
         Save the Excel file.
        ========================================================================
        """
        self._wb.save(self.path)

    def save_as(self, path: str) -> None:
        """
        ========================================================================
         Save the Excel file to a new path.
        ========================================================================
        """
        self._wb.save(path)

    def close(self) -> None:
        """
        ========================================================================
         Close the Excel file.
        ========================================================================
        """
        self._wb.close()

    @property
    def workbook(self) -> Workbook:
        """
        ========================================================================
         Get the Excel workbook.
        ========================================================================
        """
        return self._wb

    @property
    def sheet_active(self) -> Worksheet:
        """
        ========================================================================
         Get the active-sheet of the workbook.
        ========================================================================
        """
        return self._wb.active
