from openpyxl import Workbook, load_workbook
from f_file.i_0_handler import FileHandler
from f_microsoft.excel.components.sheet import Sheet


class Excel(FileHandler):
    """
    ============================================================================
     Excel file handler.
    ============================================================================
    """

    def __init__(self,
                 # Path to the Excel file to be opened
                 path: str
                 ) -> None:
        """
        ========================================================================
         Initialize the Excel file handler.
        ========================================================================
        """
        self._wb: Workbook = None
        FileHandler.__init__(self, path=path)

    def save(self) -> None:
        """
        ========================================================================
         Save the Excel file.
        ========================================================================
        """
        self._wb.save(self.path)

    def save_as(self, path: str) -> None:
        """
        ========================================================================
         Save the Excel file to a new old_path.
        ========================================================================
        """
        self._wb.save(path)

    def close(self, save: bool = True) -> None:
        """
        ========================================================================
         Close the Excel file.
        ========================================================================
        """
        if save:
            self.save()
        self._wb.close()
        
    def _create(self) -> None:
        """
        ========================================================================
         Create a new Excel file.
        ========================================================================
        """
        self._wb = Workbook()
        self.save()

    def _open(self) -> None:
        """
        ========================================================================
         Open an existing Excel file.
        ========================================================================
        """
        self._wb = load_workbook(self.path)
          
    def __getitem__(self, key: str | int) -> Sheet:
        """
        ========================================================================
         1. Get the worksheet by Name or Index.
         2. Index starts from 1.
        ========================================================================
        """
        if isinstance(key, str):
            # Get the worksheet by Name
            ws = self._wb[key]
        else:
            # Get the worksheet by Index
            ws = self._wb.worksheets[key-1]
        # Return the Sheet component
        return Sheet(sheet=ws)
