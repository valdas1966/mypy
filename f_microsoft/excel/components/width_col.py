from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


class WidthCol:
    """
    ============================================================================
     Manage the width of columns in an Excel worksheet.
    ----------------------------------------------------------------------------
     Usage:
    ----------------------------------------------------------------------------
        width_col = WidthCol(sheet)

        width = width_col[1]        # Get column width
        width_col[1] = 20           # Set column width
    ============================================================================
    """

    def __init__(self, sheet: Worksheet) -> None:
        """
        ========================================================================
         Initialize the Width-Col component.
        ========================================================================
        """
        self._sheet = sheet
        self._default = 8.43

    def __getitem__(self, col: int) -> float:
        """
        ========================================================================
         1. Get the width of a column.
         2. If the width is not set, return the default width.
        ========================================================================
        """
        letter = get_column_letter(col)
        return self._sheet.column_dimensions[letter].width or self._default

    def __setitem__(self, col: int, width: float) -> None:
        """
        ========================================================================
         Set the width of a column.
        ========================================================================
        """
        letter = get_column_letter(col)
        self._sheet.column_dimensions[letter].width = width
