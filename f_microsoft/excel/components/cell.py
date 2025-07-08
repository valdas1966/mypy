from f_core.mixins.has.row_col import HasRowCol
from f_color.rgb import RGB
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill


class Cell(HasRowCol):
    """
    ============================================================================
     1. Cell component of Microsoft Excel Worksheet.
     2. Purpose: Access Cell's properties (like value).
    ============================================================================
    """
    def __init__(self, sheet: Worksheet, row: int, col: int) -> None:
        """
        ========================================================================
         Initialize the Cell component.
        ========================================================================
        """
        # Initialize the HasRowCol mixin
        HasRowCol.__init__(self, row=row, col=col)
        # Cell to work with
        self._cell = sheet.cell(row=row, column=col)

    @property
    def value(self) -> str:
        """
        ========================================================================
         Return the value of the cell.
        ========================================================================
        """
        return self._cell.value
    
    @value.setter
    def value(self,
              # New Value
              val: str) -> None:
        """
        ========================================================================
         Set the value of the cell.
        ========================================================================
        """
        self._cell.value = val

    @property
    def background(self) -> RGB:
        """
        ========================================================================
         Return the background color of the cell.
        ========================================================================
        """
        # hex: str = self._cell.fill.start_color.rgb
        # return RGB.from_hex(hex)
        color = self._cell.fill.fgColor
        if color.rgb:  # Only use if explicitly set
            return RGB.from_hex('#' + color.rgb[-6:])

    @background.setter
    def background(self,
                   # New Background Color
                   rgb: RGB) -> None:
        """
        ========================================================================
         Set the background color of the cell.
        ========================================================================
        """
        # Fill Type of the Cell's Background (it can be only 'solid')
        fill_type = 'solid'

        # Start and End colors of the Cell's Background (it cannot be different)
        argb = rgb.to_argb()
        start_color = argb
        end_color = argb

        # Set the background color of the cell
        self._cell.fill = PatternFill(fill_type=fill_type,
                                      start_color=start_color, 
                                      end_color=end_color)
