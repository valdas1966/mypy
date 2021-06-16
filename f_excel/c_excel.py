import openpyxl as xl
from openpyxl.styles import PatternFill, Font, Alignment
import openpyxl.styles.colors as Colors
from openpyxl.styles.borders import Border, Side
from f_utils import u_file
import pandas as pd


class Excel:
    """
    ============================================================================
     Description: Class for working with Microsoft Excel.
    ============================================================================
    """

    def __init__(self, xlsx, index_ws=0):
        """
        ========================================================================
         Description: Constructor. Open Excel Workbook and Worksheet.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. filename : str (Excel Filename to open).
            2. index_ws : int (Index of the Worksheet in Excel File).
        ========================================================================
        """
        self.xlsx = xlsx
        self.wb = None
        self.ws = None
        if u_file.is_exists(xlsx):
            self.wb = xl.load_workbook(filename=xlsx)
        else:
            self.wb = xl.Workbook()
        self.ws = self.wb.worksheets[index_ws]

    def copy_worksheet(self, title_source, title_target):
        """
        ========================================================================
         Description: Make a Copy of Worksheet.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. title_source : str (Title of Source Worksheet).
            2. title_target : str (Title of New Worksheet).
        ========================================================================
         Return: Worksheet.
        ========================================================================
        """
        source = self.wb[title_source]
        target = self.wb.copy_worksheet(source)
        target.title = title_target
        return target

    def set_value(self, row, col, value):
        """
        ========================================================================
         Description: Set Value in Cell by Row and Col numbers.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. row : int
            2. col : int
            3. value : obj
        ========================================================================
        """
        assert type(row) == int
        assert type(col) == int
        assert row >= 1
        assert col >= 1
        self.ws.cell(row=row, column=col).value = value

    def set_values_col(self, row_start, col, values):
        """
        ========================================================================
         Description: Set Values to Column.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. row_start : int (First Row in Excel).
            2. col : int (Column Index).
            3. values : list (Column Values).
        ========================================================================
        """
        assert type(row_start) == int
        assert type(col) == int
        assert type(values) == list
        assert row_start > 0
        assert col > 0
        for i, val in enumerate(values):
            self.set_value(row_start + i, col, val)

    def get_value(self, row, col):
        """
        ========================================================================
         Description: Return Value in the Cell by Row and Col numbers.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. row : int
            2. col : int
        ========================================================================
         Return: obj
        ========================================================================
        """
        try:
            return self.ws.cell(row=row, column=col).value
        except Exception:
            print(f'Error in try get_value(row={row}, col={col})')
            return None

    def is_blank(self, row, col):
        """
        ========================================================================
         Description: Return True if the specified cell is blank.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. row : int
            2. col : int
        ========================================================================
         Return: bool
        ========================================================================
        """
        return self.get_value(row, col) is None

    def set_font(self, row, col, color='000000', is_bold=False):
        font = Font(color=color, bold=is_bold, vertAlign='baseline')
        self.ws.cell(row, col).font = font
        self.ws.cell(row, col).alignment = Alignment(horizontal='center',
                                                     vertical='center')

    def set_background(self, row, column, color):
        """
        ========================================================================
         Description: Set background color to specified cell.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. row : int
            2. column : int
            3. color : str
        ========================================================================
        """
        fill = Excel.color_excel(color)
        self.ws.cell(row=row, column=column).fill = fill

    def set_border(self, row, column, style='thick'):
        """
        ========================================================================
         Description: Set Block on specified Cell.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. row : int (Cell's Row).
            2. column : int (Cell's Column).
            3. style : str ('thick' | 'thin' | None).
        ========================================================================
        """
        border = None
        if style:
            border = Border(left=Side(style=style),
                            right=Side(style=style),
                            top=Side(style=style),
                            bottom=Side(style=style))
        self.ws.cell(row, column).border = border

    def clear_cells(self, row, col, row_last=None, col_last=None,
                    rows=1, cols=1, with_style=False):
        """
        ========================================================================
         Description: Clear the Cell (Value, Background, Borders).
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. row : int (First Row).
            2. col : int (First Col).
            3. row_last : int (Equals to Row on None).
            4. col_last : int (Equals to Col on None).
            5. rows : int (Amount of Rows).
            6. cols : int (Amount of Cols).
        ========================================================================
        """
        if not row_last:
            row_last = row + rows - 1
        if not col_last:
            col_last = col + cols - 1
        for r in range(row, row_last+1):
            for c in range(col, col_last+1):
                self.set_value(r, c, value=None)
                if with_style:
                    self.set_background(r, c, color='WHITE')
                    self.set_border(r, c, style='thin')

    def merge_cells(self, row, col, row_last=None, col_last=None,
                    rows=1, cols=1):
        """
        ========================================================================
         Description: Merge Cells.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. row : int (First Row).
            2. col : int (First Col).
            3. row_last : int (Equals to Row on None).
            4. col_last : int (Equals to Col on None).
            5. rows : int (Amount of Rows).
            6. cols : int (Amount of Cols).
        ========================================================================
        """
        if not row_last:
            row_last = row + rows - 1
        if not col_last:
            col_last = col + cols - 1
        self.ws.merge_cells(start_row=row, start_column=col,
                            end_row=row_last, end_column=col_last)

    def set_column_width(self, col):
        self.ws.column_dimensions[col] = 4

    def last_row(self, fr=1, col=1):
        """
        ========================================================================
         Description: Return Last-Filled Row in Particular Excel-Column.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. fr : int (Row Start).
            2. col : int (Column Index).
        ========================================================================
         Return: int
        ========================================================================
        """
        row = fr
        while not self.is_blank(row, col):
            row = row + 1
        return row-1

    def to_df(self, fr, lr, fc, lc):
        """
        ========================================================================
         Description: Convert Excel-Range into DataFrame.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. fr : int (First Row).
            2. lr : int (Last Row).
            3. fc : int (First Column).
            4. lc : int (Last Column).
        ========================================================================
         Return: DataFrame
        ========================================================================
        """
        assert type(fr) == int
        assert type(lr) == int
        assert type(fc) == int
        assert type(lc) == int
        assert fr > 0
        assert lr > 0
        assert fc > 0
        assert lc > 0
        li_col_name = list()
        li_col_values = list()
        for i_col, col in enumerate(range(fc, lc+1)):
            li_col_name.append(f'col_{i_col+1}')
            col_values = list()
            for row in range(fr, lr+1):
                col_values.append(self.get_value(row, col))
            li_col_values.append(col_values)
        d = dict(zip(li_col_name, li_col_values))
        return pd.DataFrame(d)

    def from_df(self, df, fr, fc):
        """
        ========================================================================
         Description: Paste DataFrame-Values into Excel-Range.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. df : DataFrame.
            2. fr : int (First Row).
            3. fc : int (First Col).
        ========================================================================
        """
        assert type(df) == pd.DataFrame
        assert type(fr) == int
        assert type(fc) == int
        assert fr >= 1
        assert fc >= 1
        d = df.to_dict()
        for i_col, (col_name, col_values) in enumerate(d.items()):
            col = fc + i_col
            col_values = [val for key, val in col_values.items()]
            self.set_values_col(fr, col, col_values)

    def to_linked_list(self, fr=1, fc=1):
        """
        ========================================================================
         Description: Create Linked List from Excel-Cells.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. fr : int (First Relevant Row in Excel).
            2. fc : int (First Relevant Col in Excel).
        ========================================================================
         Return: List (Row) of lists (Cols) of Str (Value).
        ========================================================================
        """
        li_rows = list()
        row = fr
        # Go Through Rows
        while not self.is_blank(row, fc):
            li_cols = list()
            col = fc
            # Go Through Columns
            while not self.is_blank(row, col):
                value = self.get_value(row, col)
                li_cols.append(value)
                col += 1
            li_rows.append(li_cols)
            row += 1
        return li_rows

    def save_as(self, xlsx):
        """
        ========================================================================
         Description: Save Excel-File as other Name.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. xlsx : str (Other Name for Excel-File).
        ========================================================================
        """
        self.xlsx = xlsx
        self.wb.save(self.xlsx)

    def close(self):
        """
        ========================================================================
         Description: Save and Close the Workbook.
        ========================================================================
        """
        self.wb.save(self.xlsx)
        self.wb.close()

    @staticmethod
    def color_excel(name_color):
        """
        ========================================================================
         Description: Get Color's Name and return its Excel PatternFill.
        ========================================================================
         Arguments:
        ------------------------------------------------------------------------
            1. name_color : str (Color's Name).
        ========================================================================
        """
        enum = {'BLACK': 0, 'WHITE': 1, 'YELLOW': 51, 'GREEN': 57, 'RED': 60,
                'GRAY': 55}
        color = Colors.COLOR_INDEX[enum[name_color]]
        return PatternFill(start_color=color,
                           end_color=color,
                           fill_type='solid')

